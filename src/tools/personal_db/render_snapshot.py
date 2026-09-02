#!/usr/bin/env python3
"""
render_snapshot.py — generate the xlsx snapshot views, one per domain.

Two xlsx snapshots, from two different sources of truth:
  books        -> Zotero          -> <snapshot_base>/library_snapshots/library.xlsx
  applications -> personal.db     -> <snapshot_base>/applications_snapshots/applications.xlsx

Books live in Zotero: personal.db holds no books domain, so the
library snapshot is read through tools/zotero/zotero_export.py, which copies
zotero.sqlite and reads the copy — safe with Zotero running. The `domains`
registry still names the output file; `source` in SPECS says where the rows
come from.

Each xlsx is a generated, always-sorted view used as a visual check and
mistake-finding aid, never an input. It is overwritten in place — the live file
is never dated, so there is only ever one "current" answer.

The books domain additionally keeps a dated HISTORY, because the library's past
states are wanted in their own right (the 2016 export, the migration off Google
Sheets). After rendering, the live file is copied into
library_snapshots/archive/library_YYYY-MM-DD.xlsx and the series is pruned to a
grandfather-father-son policy — see snapshot_archive.py, which owns that
behaviour and explains the retention. One file per day maximum, so running this
repeatedly in an afternoon does not pile up.

Usage:
  PERSONAL_DB_DIR=... python3 render_snapshot.py [--domain applications|learning|books|all]
                                                 [--no-archive]

Prereq: openpyxl
"""

import shutil
import sys
import tempfile
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.table import Table, TableStyleInfo
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("ERROR: openpyxl not installed — pip install openpyxl --break-system-packages")

sys.path.insert(0, str(Path(__file__).resolve().parent))
sys.path.insert(0, str(Path(__file__).resolve().parents[1] / "zotero"))
import db_common as dbc  # noqa: E402
import snapshot_archive as sa  # noqa: E402
import zotero_export as ze  # noqa: E402

HEADER_FILL = PatternFill("solid", fgColor="4F81BD")
HEADER_FONT = Font(bold=True, color="FFFFFF")
ALT_FILL = PatternFill("solid", fgColor="DCE6F1")

# ── Render specs ─────────────────────────────────────────────────────────────
# Per domain: the snapshot subdir (under snapshot_base), sheet title, ordered
# columns (header, field, width), where the rows come from (`sql` against
# personal.db, or `source: "zotero"`), a table name, and the Stats-tab rows
# (label, live-Excel-formula). `books` keeps the library's full layout and all
# 14 Stats metrics, which are Excel formulas over the sheet — so they follow
# the data wherever it is read from.
SPECS = {
    "applications": {
        "subdir": "applications_snapshots",
        "source": "personal_db",
        "sheet": "Applications",
        "table": "Applications",
        "columns": [
            ("Company", "company", 24), ("Role", "role", 34),
            ("Fit Notes", "fit_notes", 40), ("Fit Verdict", "fit_verdict", 12),
            ("Gaps", "gaps", 22), ("Location", "location", 14),
            ("ATS Platform", "ats_platform", 14), ("Date Evaluated", "date_evaluated", 14),
            ("Cover Letter", "cover_letter", 12), ("Status", "status", 16),
            ("Contact", "contact", 16), ("Referral", "referral", 12),
            ("JD Link", "jd_link", 30), ("Year", "year", 8),
        ],
        "sql": """SELECT company, role, fit_notes, fit_verdict, gaps, location,
                         ats_platform, date_evaluated, cover_letter, status,
                         contact, referral, jd_link, year
                  FROM applications
                  ORDER BY year DESC, LOWER(company)""",
        "stats": [
            ("Total Applications", "=COUNTA(Applications!A2:A100000)"),
            ("Applied", '=COUNTIF(Applications!J:J,"Applied*")'),
            ("Interviewing", '=COUNTIF(Applications!J:J,"Interview*")'),
            ("Rejected", '=COUNTIF(Applications!J:J,"Rejected*")'),
            ("Pending", '=COUNTIF(Applications!J:J,"Pending*")'),
            ("With Cover Letter", '=COUNTIF(Applications!I:I,"Yes")'),
        ],
    },
    "learning": {
        "subdir": "learning_snapshots",
        "source": "personal_db",
        "sheet": "Learning",
        "table": "Learning",
        "columns": [
            ("Course", "course", 26), ("Lesson", "lesson", 8),
            ("Kind", "kind", 12), ("Item", "item", 22),
            ("Score", "score", 12), ("Recorded", "recorded_at", 20),
        ],
        "sql": """SELECT course, lesson, kind, item, score, recorded_at
                  FROM learning_progress
                  ORDER BY LOWER(course), lesson, kind, item""",
        # C = Kind
        "stats": [
            ("Lessons Opened", '=COUNTIF(Learning!C:C,"opened")'),
            ("Lessons Read", '=COUNTIF(Learning!C:C,"reading")'),
            ("Quizzes Answered", '=COUNTIF(Learning!C:C,"quiz")'),
            ("Exercises Done", '=COUNTIF(Learning!C:C,"exercise")'),
        ],
    },
    "books": {
        "subdir": "library_snapshots",
        "source": "zotero",
        # Keep the dated series (snapshot_archive.py). Books only: the library
        # is a collection whose history is interesting, applications are a
        # working list whose past states are not.
        "archive": True,
        "sheet": "My Library",
        "table": "MyLibrary",
        "columns": [
            ("Author", "author", 30), ("Title", "title", 45),
            ("Publisher", "publisher", 25), ("Edition", "edition", 15),
            ("Signed", "signed", 7), ("Genre", "genre", 20),
            ("Pub Date", "pub_date", 9), ("Page Count", "page_count", 9),
            ("Price Paid", "price_paid", 9), ("Read", "read", 6),
            ("Shelved", "shelved", 8),
        ],
        # H=Page Count, J=Read, K=Shelved
        "stats": [
            ("Total Pages Read", "=SUMIFS('My Library'!H:H,'My Library'!J:J,\"Y\")"),
            ("Total Titles Read", "=COUNTIF('My Library'!J:J,\"Y\")"),
            ("Percent of Logged Read",
             "=COUNTIF('My Library'!J:J,\"Y\")/(COUNTIF('My Library'!J:J,\"Y\")+COUNTIF('My Library'!J:J,\"N\"))"),
            ("Avg Length, Read Titles", "=AVERAGEIF('My Library'!J:J,\"Y\",'My Library'!H:H)"),
            ("Avg Length, Read+Shelved",
             "=AVERAGEIFS('My Library'!H:H,'My Library'!J:J,\"Y\",'My Library'!K:K,\"Y\")"),
            ("Avg Length, Unread Titles", "=AVERAGEIF('My Library'!J:J,\"N\",'My Library'!H:H)"),
            ("Shelved Pages Read", "=SUMIFS('My Library'!H:H,'My Library'!J:J,\"Y\",'My Library'!K:K,\"Y\")"),
            ("Shelved Titles Read", "=COUNTIFS('My Library'!J:J,\"Y\",'My Library'!K:K,\"Y\")"),
            ("Unshelved Pages Read", "=SUMIFS('My Library'!H:H,'My Library'!J:J,\"Y\",'My Library'!K:K,\"N\")"),
            ("Unshelved Titles Read", "=COUNTIFS('My Library'!J:J,\"Y\",'My Library'!K:K,\"N\")"),
            ("Total Shelved Pages", "=SUMIF('My Library'!K:K,\"Y\",'My Library'!H:H)"),
            ("Total Shelved Titles", "=COUNTIF('My Library'!K:K,\"Y\")"),
            ("Shelved Pages Not Read", "=SUMIFS('My Library'!H:H,'My Library'!K:K,\"Y\",'My Library'!J:J,\"N\")"),
            ("Shelved Titles Not Read", "=COUNTIFS('My Library'!K:K,\"Y\",'My Library'!J:J,\"N\")"),
        ],
    },
}


def _num(value):
    """'$5.00' -> 5.0, '1963' -> 1963, anything else unchanged. Zotero stores
    every field as text; the snapshot's numeric columns were numbers before the
    move and the Stats formulas sum them, so they have to stay numbers."""
    if not isinstance(value, str):
        return value
    text = value.strip().lstrip("$").replace(",", "")
    try:
        return int(text)
    except ValueError:
        pass
    try:
        return float(text)
    except ValueError:
        return value


def zotero_rows(spec) -> list[tuple]:
    """The library, in the snapshot's column order. Sorted by author then title,
    which is the order zotero_export already returns."""
    fields = [f for _, f, _ in spec["columns"]]
    numeric = {"pub_date", "page_count", "price_paid"}
    return [tuple(_num(r[f]) if f in numeric else r[f] for f in fields)
            for r in ze.read_books()]


def render_domain(domain: str, conn, out_path: Path) -> int:
    spec = SPECS[domain]
    out_path.parent.mkdir(parents=True, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = spec["sheet"]

    headers = [c[0] for c in spec["columns"]]
    ws.append(headers)
    for i, (_, _, w) in enumerate(spec["columns"], 1):
        cell = ws.cell(row=1, column=i)
        cell.font, cell.fill = HEADER_FONT, HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    rows = (zotero_rows(spec) if spec["source"] == "zotero"
            else conn.execute(spec["sql"]).fetchall())
    for ri, row in enumerate(rows, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)
        if ri % 2 == 0:
            for ci in range(1, len(spec["columns"]) + 1):
                ws.cell(row=ri, column=ci).fill = ALT_FILL

    last_col = get_column_letter(len(spec["columns"]))
    tbl = Table(displayName=spec["table"], ref=f"A1:{last_col}{max(ws.max_row, 2)}")
    tbl.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    ws.add_table(tbl)

    if spec.get("stats"):
        wss = wb.create_sheet("Stats")
        wss.append(["Metric", "Value"])
        for c in (1, 2):
            wss.cell(1, c).font, wss.cell(1, c).fill = HEADER_FONT, HEADER_FILL
        wss.column_dimensions["A"].width = 32
        wss.column_dimensions["B"].width = 18
        for label, formula in spec["stats"]:
            r = wss.max_row + 1
            wss.cell(r, 1, label)
            wss.cell(r, 2, formula)

    tmp_out = Path(tempfile.gettempdir()) / f"personal_snap_{domain}.xlsx"
    wb.save(str(tmp_out))
    shutil.copy2(str(tmp_out), str(out_path))
    tmp_out.unlink(missing_ok=True)
    return len(rows)


def main() -> None:
    target = "all"
    for i, a in enumerate(sys.argv):
        if a == "--domain" and i + 1 < len(sys.argv):
            target = sys.argv[i + 1]
    keep_history = "--no-archive" not in sys.argv

    conn = dbc.connect()
    domains = {r["name"]: r for r in conn.execute(
        "SELECT name, snapshot_file FROM domains WHERE active=1")}
    base = dbc.snapshot_base()
    todo = list(domains) if target == "all" else [target]

    for d in todo:
        if d not in SPECS:
            print(f"  ⚠ no render spec for domain '{d}' — skipped")
            continue
        out = base / SPECS[d]["subdir"] / domains[d]["snapshot_file"]
        n = render_domain(d, conn, out)
        print(f"[render_snapshot] {d:14s} {n:5d} rows -> {out}")
        if keep_history and SPECS[d].get("archive"):
            dest, dropped = sa.archive_and_prune(out)
            print(f"[render_snapshot] {'':14s} history -> {dest.name}"
                  f" ({len(dropped)} pruned)")
    conn.close()


if __name__ == "__main__":
    main()
